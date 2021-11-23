from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pprint import pprint
from scrap import getProduct

scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)
client = gspread.authorize(creds)
sheet = client.open("AnalyticaHouse").sheet1  # Open the spreadhseet
fields = [ "URL", "Product Name", "Availability", "Offer", "Sale Price", "Product Price", "Product Code" ]

#append given product to the given google sheet.
def appendProduct( product, sheet ):
    productRow = [ 
                    product.url,
                    product.productName,
                    "%" + str( product.availability ),
                    product.offer,
                    product.salePrice,
                    product.productPrice,
                    product.productCode
                ]
    #print( productRow )
    sheet.append_row( productRow )



#open URL's.xlsx file. Fill all urls to allUrls[] list
wb = load_workbook(filename = 'URL\'s.xlsx')
ws = wb.active
allUrls = []
i = 1
currentUrl =  ws['A1'].value
while( currentUrl != None ):
    #print( currentUrl )
    allUrls.append( "https://www.markastok.com" + currentUrl )
    i += 1
    currentUrl = ws[ 'A' + str(i) ].value


#retrieve and send all product informations to the google sheet.
for i in range( 0, len( allUrls ) ):
    currentProduct = getProduct( allUrls[i] )
    currentType = currentProduct.productType
    if( currentType == "available-product" or currentType == "no-stock" ):
        appendProduct( currentProduct, sheet )




